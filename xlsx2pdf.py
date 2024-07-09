import pandas as pd
from spire.xls import *
from spire.xls.common import *
import pythoncom
import openpyxl


def xlsx2pdf(upload_path, result_path, file_name):
    pythoncom.CoInitialize()

    # 시트 이름 리스트를 불러온다.
    names = openpyxl.load_workbook(upload_path).sheetnames
    
    # 복수의 스프레드시트를 하나의 시트로 합친다.
    df = pd.DataFrame({})
    for name in names:
        temp_df = pd.read_excel(upload_path, sheet_name=name)
        # temp_df['시트이름'] = name #시트이름의 날짜정보를 새로운 열로 생성
        df= pd.concat([df,temp_df])
    
    #파일로 저장
    merged_xlsx = f'uploads/{file_name}_merged.xlsx'
    df.to_excel(merged_xlsx, index=False)
    
    # workbook을 생성하고 merge된 파일을 로드한다.
    workbook = Workbook()
    workbook.LoadFromFile(merged_xlsx)

    # 통합된 엑셀파일이기 때문에 첫 번째 시트를 선택
    sheet= workbook.Worksheets[0]
    result_file = os.path.abspath(os.path.join(result_path, file_name)+ '.pdf')
    sheet.SaveToPdf(result_file)

    workbook.Dispose()

